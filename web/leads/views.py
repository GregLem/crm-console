from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from .models import Lead
from .forms import LeadForm
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def lead_list(request):
    """Обычный список (оставим как есть)"""
    leads = Lead.objects.all().order_by('-created_at')
    return render(request, 'leads/lead_list.html', {'leads': leads})

def add_lead(request):
    """Добавление заявки"""
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lead_list')
    else:
        form = LeadForm()
    return render(request, 'leads/add_lead.html', {'form': form})

def kanban_board(request):
    """Канбан-доска"""
    statuses = Lead.STATUS_CHOICES
    leads_by_status = {}
    
    for status_value, status_label in statuses:
        leads_by_status[status_value] = {
            'label': status_label,
            'leads': Lead.objects.filter(status=status_value).order_by('-created_at')
        }
    
    context = {
        'leads_by_status': leads_by_status,
        'statuses': statuses,
    }
    return render(request, 'leads/kanban_board.html', context)

def update_lead_status(request):
    """API для перетаскивания карточек"""
    if request.method == 'POST':
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        new_status = data.get('status')
        
        try:
            lead = Lead.objects.get(id=lead_id)
            lead.status = new_status
            lead.save()
            return JsonResponse({'success': True})
        except Lead.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Лид не найден'})
    
    return JsonResponse({'success': False})

def analytics_dashboard(request):
    """Аналитика с графиками"""
    today = datetime.now().date()
    week_ago = today - timedelta(days=6)
    
    dates = []
    leads_count = []
    
    for i in range(7):
        current_date = week_ago + timedelta(days=i)
        dates.append(current_date.strftime('%d.%m'))
        count = Lead.objects.filter(created_at__date=current_date).count()
        leads_count.append(count)
    
    # Статистика по статусам
    status_stats = {}
    for status_value, status_label in Lead.STATUS_CHOICES:
        status_stats[status_label] = Lead.objects.filter(status=status_value).count()
    
    context = {
        'dates': json.dumps(dates),
        'leads_count': json.dumps(leads_count),
        'status_labels': json.dumps(list(status_stats.keys())),
        'status_data': json.dumps(list(status_stats.values())),
        'total_leads': Lead.objects.count(),
        'new_leads_today': Lead.objects.filter(created_at__date=today).count(),
    }
    return render(request, 'leads/analytics.html', context)

def export_leads_excel(request):
    """Выгрузка в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заявки'
    
    # Заголовки
    headers = ['ID', 'Имя', 'Телефон', 'Услуга', 'Источник', 'Статус', 'Дата создания', 'Комментарий']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4dabf7', end_color='4dabf7', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    leads = Lead.objects.all().order_by('-created_at')
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead.id)
        ws.cell(row=row, column=2, value=lead.name)
        ws.cell(row=row, column=3, value=lead.phone)
        ws.cell(row=row, column=4, value=lead.service)
        ws.cell(row=row, column=5, value=lead.source)
        ws.cell(row=row, column=6, value=lead.get_status_display())
        ws.cell(row=row, column=7, value=lead.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=8, value=lead.comment)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="leads_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response